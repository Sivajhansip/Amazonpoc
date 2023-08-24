import openpyxl


class Excel:
    def __init__(self, path):
        self.path = path
        self.book = openpyxl.load_workbook(path)

    def row_count(self):
        sheet = self.book.active
        return sheet.max_row

    def col_count(self):
        sheet = self.book.active
        return sheet.max_column

    def read_data(self, rowNum, colNum):
        sheet = self.book.active
        return sheet.cell(row=rowNum, column=colNum).value

    def write_data(self, summary, result):
        sheet = self.book.active
        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1).value = next_row - 1
        sheet.cell(row=next_row, column=2).value = summary
        sheet.cell(row=next_row, column=3).value = result
        self.book.save(self.path)
